import re
import sqlite3
from datetime import datetime

import openpyxl

DB_PATH = "vendor_contracts.db"
EXCEL_PATH = "IT_Service_Contract_Details_Fy25-26.xlsx"
SHEET_NAME = "Fy25-26"
HEADER_ROW = 3
DATA_START_ROW = 4

COL_TYPE = 2
COL_DETAILS = 3
COL_VENDOR = 4
COL_MASTER_CONTRACT = 5
COL_PO = 6
COL_START_DATE = 7
COL_END_DATE = 8
COL_AMOUNT = 9
COL_STATUS = 10
COL_PR_NO = 11
COL_REMARKS_1 = 12
COL_REMARKS_2 = 13
COL_DUE_DATE = 14
COL_OLD_PO = 15

MIN_REASONABLE_YEAR = 2015
MAX_REASONABLE_YEAR = 2045


def clean_text(value):
   
    if value is None:
        return None
    text = " ".join(str(value).split())
    return text if text else None


def parse_date(value):
    
    if value is None:
        return None, None

    if isinstance(value, datetime):
        dt = value
    else:
        text = str(value).strip()
        if not text:
            return None, None
        dt = None
        for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                dt = datetime.strptime(text, fmt)
                break
            except ValueError:
                continue
        if dt is None:
            return None, f"could not parse date '{text}', left blank"

    if not (MIN_REASONABLE_YEAR <= dt.year <= MAX_REASONABLE_YEAR):
        return None, f"date {dt.date()} looks like a typo (year out of range), left blank"

    return dt.date().isoformat(), None


def parse_po_number(value):
    
    if value is None:
        return None, None

    if isinstance(value, (int, float)):
        return str(int(value)), None

    text = str(value).strip()
    if not text or text.lower() == "new":
        return None, None

    numbers = re.findall(r"\d+", text)
    if not numbers:
        return None, f"PO field '{text}' had no digits, left blank"

    if len(numbers) > 1:
        return numbers[0], f"additional PO number(s) in the same cell: {', '.join(numbers[1:])}"

    return numbers[0], None


def to_text_number(value):
    """PR numbers / OldPO sometimes come through as floats (e.g. 5000038038.0)."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return str(int(value))
    return clean_text(value)


def build_vendor_lookup(conn):
    rows = conn.execute("SELECT vendor_id, name FROM vendors").fetchall()
    return {name: vid for vid, name in rows}


def migrate():
    conn = sqlite3.connect(DB_PATH)

    existing_count = conn.execute("SELECT COUNT(*) FROM contracts").fetchone()[0]
    if existing_count > 0:
        print(f"contracts table already has {existing_count} row(s).")
        print("This script doesn't check for duplicates, so running it again")
        print("would double up your data. Delete vendor_contracts.db and rebuild")
        print("with create_db.py + migrate_vendors.py first if you want a clean re-run.")
        conn.close()
        return

    vendor_lookup = build_vendor_lookup(conn)

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    sheet = wb[SHEET_NAME]

    inserted = 0
    skipped_no_vendor = []
    skipped_vendor_not_found = []
    notes_log = []

    for row_num in range(DATA_START_ROW, sheet.max_row + 1):
        def get(col):
            return sheet.cell(row=row_num, column=col).value

        vendor_name = clean_text(get(COL_VENDOR))
        details = clean_text(get(COL_DETAILS))

        if vendor_name is None:
            if details or get(COL_TYPE):
                skipped_no_vendor.append((row_num, details))
            continue

        vendor_id = vendor_lookup.get(vendor_name)
        if vendor_id is None:
            skipped_vendor_not_found.append((row_num, vendor_name))
            continue

        contract_type = clean_text(get(COL_TYPE))
        master_note = clean_text(get(COL_MASTER_CONTRACT))
        amount = get(COL_AMOUNT)
        status = clean_text(get(COL_STATUS))
        pr_no = to_text_number(get(COL_PR_NO))

        po_number, po_note = parse_po_number(get(COL_PO))
        start_date, start_note = parse_date(get(COL_START_DATE))
        end_date, end_note = parse_date(get(COL_END_DATE))
        due_date, due_note = parse_date(get(COL_DUE_DATE))

        remarks_parts = [clean_text(get(COL_REMARKS_1)), clean_text(get(COL_REMARKS_2))]
        old_po_text = to_text_number(get(COL_OLD_PO))
        if old_po_text:
            remarks_parts.append(f"(previous PO on record: {old_po_text})")
        if po_note:
            remarks_parts.append(f"(PO note: {po_note})")
        remarks = " | ".join(p for p in remarks_parts if p) or None

        for label, note in (("start_date", start_note), ("end_date", end_note), ("due_date", due_note)):
            if note:
                notes_log.append(f"row {row_num} ({vendor_name}): {label} - {note}")

        conn.execute(
            """
            INSERT INTO contracts
                (vendor_id, contract_type, details, master_contract_note, po_number,
                 start_date, due_date, end_date, yearly_amount, procurement_status,
                 pr_number, remarks)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (vendor_id, contract_type, details, master_note, po_number,
             start_date, due_date, end_date, amount, status, pr_no, remarks),
        )
        inserted += 1

    conn.commit()
    conn.close()

    print(f"Inserted {inserted} contracts.\n")

    if skipped_no_vendor:
        print(f"Skipped {len(skipped_no_vendor)} row(s) with no vendor name (nothing to link the contract to):")
        for row_num, details in skipped_no_vendor:
            print(f"  - row {row_num}: {details!r}")
        print()

    if skipped_vendor_not_found:
        print(f"Skipped {len(skipped_vendor_not_found)} row(s) whose vendor name wasn't found in the vendors table:")
        for row_num, name in skipped_vendor_not_found:
            print(f"  - row {row_num}: {name!r}")
        print("  (re-run migrate_vendors.py first if these are genuinely missing)")
        print()

    if notes_log:
        print(f"{len(notes_log)} date(s)/field(s) worth double-checking by hand:")
        for note in notes_log:
            print(f"  - {note}")


if __name__ == "__main__":
    migrate()