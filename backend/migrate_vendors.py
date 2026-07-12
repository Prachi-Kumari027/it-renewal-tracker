import sqlite3
import openpyxl

DB_PATH = "vendor_contracts.db"
EXCEL_PATH = "IT_Service_Contract_Details_Fy25-26.xlsx"
SHEET_NAME = "Fy25-26"
HEADER_ROW = 3          
DATA_START_ROW = 4      
VENDOR_COLUMN = 4      


def read_vendor_names():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    sheet = wb[SHEET_NAME]

    header_value = sheet.cell(row=HEADER_ROW, column=VENDOR_COLUMN).value
    if header_value != "Vendor":
        raise ValueError(
            f"Expected 'Vendor' header at row {HEADER_ROW}, column {VENDOR_COLUMN}, "
            f"but found {header_value!r} instead. Check if the sheet layout changed."
        )

    raw_names = []
    for row in range(DATA_START_ROW, sheet.max_row + 1):
        value = sheet.cell(row=row, column=VENDOR_COLUMN).value
        if value is not None and str(value).strip() != "":
            raw_names.append(str(value))

    messy = [n for n in raw_names if "\n" in n]
    if messy:
        print("Heads up - these vendor cells contain a line break and look messy.")
        print("They'll still be migrated (line breaks replaced with spaces),")
        print("but you may want to clean them up by hand later:")
        for m in messy:
            print("  -", repr(m))
        print()

    cleaned = [" ".join(n.split()) for n in raw_names] 
    unique_names = sorted(set(cleaned))
    return unique_names


def insert_vendors(names):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    inserted = 0
    for name in names:
        cur.execute("INSERT OR IGNORE INTO vendors (name) VALUES (?)", (name,))
        if cur.rowcount:
            inserted += 1

    conn.commit()
    total = cur.execute("SELECT COUNT(*) FROM vendors").fetchone()[0]
    conn.close()
    return inserted, total


if __name__ == "__main__":
    names = read_vendor_names()
    print(f"Found {len(names)} unique vendor names in the Excel sheet.\n")

    inserted, total = insert_vendors(names)
    print(f"Inserted {inserted} new vendor(s). Table now has {total} vendor(s) total.")